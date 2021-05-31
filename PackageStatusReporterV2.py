""" TODO
-Reduce memory usage
-Add package tracking status
-Closed but not shipped report?
"""

### Imports
## Default libraries
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import datetime as datetime_module
from datetime import datetime
import csv
import json
import sys
import traceback
import os
import shutil

## 3rd Party Libraries (installable through pip)
import openpyxl as pyxl

## Distributed with PackageStatusReporter
import Google_API as gapi

### Classes

### Functions
## I/O Functions

def load_config():
	try:
		with open("resources/CONFIG.json", "r") as file:
			data = json.load(file)
		return data
	except:
		show_error("ERROR", "Unable to load necessary file ./resources/CONFIG.json.\nFile missing?")
		log_error()

def save_config():
	try:
		global config_data
		with open("resources/CONFIG.json", "w") as file:
			json.dump(config_data, file, indent = 4)
	except:
		log_error()

def load_database():
	global config_data
	global database_data
	try:
		#Has a pre-set structure, so no parsing
		database_data = []
		with open(config_data.get("database_path"), "r", encoding="utf-8") as file:
			csv_reader = csv.reader(file, delimiter = ",")
			for row in csv_reader:
				if is_not_empty(row):
					database_data.append(right_pad_list(row, config_data["database_row_length"], ""))
		database_status_label.config(text = str(len(database_data)) + " entries in database")
	except:
		show_error("ERROR", "Unable to load database!\nMake sure \"database_path\" in config.json points to database\nCheck if database is open in another program")
		log_error()

def save_database(data_in):
	try:
		global config_data
		global database_data
		activity_log(f"DATABASE saving {len(database_data)} entries. . .")

		backup_database()
		with open(config_data["database_path"], "w", newline = "", encoding="utf-8") as file:
			csv_writer = csv.writer(file)
			for row in data_in:
				#print("OUT:", row)
				csv_writer.writerow(row)
		database_status_label.config(text = str(len(database_data)) + " entries in database")
		show_info("SUCCESS", "Database successfully updated!")
		activity_log(f"DATABASE saved")
	except:
		show_error("ERROR", "Unable to save database!\nMake sure database is not open in another program")
		activity_log(f"ERROR Unable to save Database!")
		log_error()

def activity_log(string_in: str):
	print(string_in)
	if not os.path.exists("resources/Activity_Log.txt"):
		file = open("resources/Activity_Log.txt", "w")
		file.write("Activity Log")
		file.close()

	with open("resources/Activity_Log.txt", "a") as file:
		file.write("\n" + str(datetime.strftime(datetime.now(), "%m-%d %H:%M:%S")) + "\t" + string_in)

def backup_database():
	global config_data
	global database_data
	source_filepath = config_data["database_path"]
	destination_filepath = (source_filepath[:-4] + " " + str(datetime.strftime(datetime.now(), "%m-%d %H.%M.%S")) + ".csv")
	config_data["last_backup_path"] = destination_filepath
	activity_log(f"DATABASE Backing up Database to {destination_filepath}. . .")
	save_config()

	shutil.copyfile(source_filepath, destination_filepath)
	clean_old_backups()
	activity_log("DATABASE Successfully backed up!")

def restore_database():
	global config_data
	global database_data
	source_filepath = config_data["last_backup_path"]
	destination_filepath = config_data["database_path"]
	activity_log(f"DATABASE Restoring Database from {source_filepath}. . .")

	shutil.copyfile(source_filepath, destination_filepath)
	activity_log("DATABASE Successfully restored backup!")

def clean_old_backups():
	global config_data
	files = os.listdir("resources")
	database_filepath = config_data["database_path"].split("/")[-1]
	backup_filepath = config_data["last_backup_path"].split("/")[-1]
	for file in files:
		#print(file, file == database_filepath, file == backup_filepath)
		if file.endswith(".csv") and file.startswith("DATABASE"):
			if file != database_filepath and file != backup_filepath:
				try:
					activity_log(f"DATABASE Deleting outdated backup {file}")
					os.remove("resources/" + file)
				except:
					activity_log(f"DATABASE Unable to delete {file}")

## Scanning Functions

def scans_list_button():
	try:
		#Get scans from textbox, splits in to list
		scans_list = scans_list_textbox.get("1.0", tk.END).split("\n")

		#Removes duplicates and mis-scanned entries
		good_scans = []
		for scan in scans_list:
			if (len(scan) == 6 or len(scan) == 8) and is_only_numeric(scan) and scan not in good_scans:
				good_scans.append(scan)
		if len(good_scans) > 0:
			match_scans(good_scans)
		else:
			show_error("ERROR", "No valid scans found!")
	except: pass

def match_scans(scans_list):
	#Loads database, compares scans to Order/Batch IDs, outputs matches
	global database_data
	global config_data
	load_database()
	out_data = []
	matches_list = []

	for i, row in enumerate(database_data):
		out_row = row
		for scan in scans_list:
			if row[config_data["database_OrderID_index"]] == scan or row[config_data["database_BatchID_index"]] == scan: #Matches either to pick ticket scan
				matches_list.append(row[2])
				row[config_data["database_row_length"] - 2] = "SHIPPED" #Row[9] <- StalcoStatus #TODO Shrink
				#Possibly temporary date specification for uploads
				if scans_date_entrybox.get() == "":
					row[config_data["database_row_length"] - 1] = str(datetime.strftime(datetime.now(), "%Y-%m-%d"))[0:10] #Row[10] <- StalcoStatusTimestamp #TODO Shrink
				else:
					row[config_data["database_row_length"] - 1] = scans_date_entrybox.get()
				database_data[i] = row #Warning: In-loop element modification. Bad practice, but good for speed

	if len(matches_list) < len(scans_list):
		show_error("ERROR", f"{len(matches_list)} matches found, {len(scans_list)} expected!\nPlease update database with the entries listed in /resources/Activity_Log.txt")
		activity_log(f"ERROR loading {scans_file_entrybox.get()} {len(matches_list)} matches found, {len(scans_list)} expected. Missing Entries:")
		for scan in scans_list:
			if scan not in matches_list:
				activity_log(scan)
	else:
		show_info("Success!", f"All {len(matches_list)} scans matched database entries!\nPlease wait while we update the database. . .")
		activity_log(f"SUCCESS loading {scans_file_entrybox.get()} {len(matches_list)} matches")

	save_database(database_data)

def scans_browse_filesystem():
	filename = filedialog.askopenfilename()
	scans_file_entrybox.delete(0, tk.END) #Clears the entrybox
	if filename:
		if filename[-4:] == ".csv" or filename[-5:] == ".xlsx":
			scans_file_entrybox.insert(tk.END, filename) #Sets entrybox to filename
		else:
			show_error("ERROR", "Invalid file!\nFile must be .csv or .xlsx")

def load_scans():
	try:
		filename = scans_file_entrybox.get()

		if filename[-4:] == ".csv":
			data_out = []
			with open(filename, "r") as file:
				csv_reader = csv.reader(file, delimiter = ",")
				for row in csv_reader:
					if is_not_empty(row):
						data_out.append(row)
				header_line = data_out.pop(0)
			parse_scans(header_line, data_out)

		elif filename[-5:] == ".xlsx":
			data_out = []
			worksheet = pyxl.load_workbook(filename).active
			for row in worksheet.values:
				if is_not_empty(row):
					data_out.append(row)
			header_line = data_out.pop(0)
			parse_scans(header_line, data_out)
			
		else:
			show_error("ERROR", "Unable to load file!\nMake sure file ends in .csv or .xlsx")
	except:
		show_error("ERROR", "Unable to load file!\nIs it open in another program?")
		activity_log(f"ERROR loading {scans_file_entrybox.get()}! Misc. Error")
		log_error()

def parse_scans(header_line, data_in):
	global config_data
	batch_column_index = -1
	for i, cell in enumerate(header_line):
		if cell != None:
			if fuzzy_match(cell, config_data["scans_sheet_batches_headers"]):
				batch_column_index = i
	if batch_column_index != -1: #If a match was found
		good_scans = []
		for row in data_in:
			scan = str(row[batch_column_index])
			if (len(row) == 6 or len(scan) == 8) and is_only_numeric(scan) and scan not in good_scans:
				good_scans.append(scan)
		match_scans(good_scans)
	else:
		show_error("ERROR", "Unable to find \"Batches\" in the header line of uploaded file")
		activity_log(f"ERROR loading {scans_file_entrybox.get()} Missing \"Batches\" in header")

def upload_scans_folder():
	scans_list = os.listdir("scans")
	set_verbose(False)
	for scan in scans_list:
		try:
			filename = "scans/" + scan
			print("BULKSCAN Uploading", filename)

			#Tries to read the filename and use it as the scan date
			file_date = ""
			try:
				file_date = str(datetime.strptime(scan.split(" ")[0], "%Y-%m-%d"))[0:10]
				print("BULKSCAN Parsed filename as", file_date)
			except:
				file_date = str(datetime.strftime(datetime.now(), "%Y-%m-%d"))[0:10]
				print("BULKSCAN Unable to parse filename. Defaulting to", file_date)
			scans_date_entrybox.delete(0, tk.END)
			scans_date_entrybox.insert(tk.END, file_date)

			if filename[-4:] == ".csv":
				data_out = []
				with open(filename, "r") as file:
					csv_reader = csv.reader(file, delimiter = ",")
					for row in csv_reader:
						if is_not_empty(row):
							data_out.append(row)
					header_line = data_out.pop(0)
				parse_scans(header_line, data_out)

			elif filename[-5:] == ".xlsx":
				data_out = []
				worksheet = pyxl.load_workbook(filename).active
				for row in worksheet.values:
					if is_not_empty(row):
						data_out.append(row)
				header_line = data_out.pop(0)
				parse_scans(header_line, data_out)
				
			else:
				show_error("ERROR", "Unable to load file!\nMake sure file ends in .csv or .xlsx")
			print("BULKSCAN Finished!")
			
		except:
			show_error("ERROR", "Unable to load file!\nIs it open in another program?")
			activity_log(f"ERROR loading {scans_file_entrybox.get()}! Misc. Error")
			log_error()

## Database Updating Functions

def database_browse_filesystem():
	filename = filedialog.askopenfilename()
	database_file_entrybox.delete(0, tk.END) #Clears the entrybox
	if filename:
		if filename[-4:] == ".csv" or filename[-5:] == ".xlsx":
			database_file_entrybox.insert(tk.END, filename) #Sets entrybox to filename
		else:
			show_error("ERROR", "Invalid file!\nFile must be .csv or .xlsx")

def load_new_entries():
	try:
		filename = database_file_entrybox.get()
		if filename[-4:] == ".csv":
			data_out = []
			with open(filename, "r", encoding = "utf-8") as file:
				csv_reader = csv.reader(file, delimiter = ",")
				for row in csv_reader:
					if is_not_empty(row):
						data_out.append(row)
				header_line = data_out.pop(0)
			if len(data_out) > 0:
				parse_new_entries(header_line, data_out)
			else:
				show_error("ERROR", "No valid scans found!")
				activity_log(f"ERROR No valid scans found in {filename}")

		elif filename[-5:] == ".xlsx":
			data_out = []
			worksheet = pyxl.load_workbook(filename).active
			for row in worksheet.values:
				if is_not_empty(row):
					data_out.append(row)
			header_line = data_out.pop(0)
			if len(data_out) > 0:
				parse_new_entries(header_line, data_out)
			else:
				show_error("ERROR", "No valid scans found!")
				activity_log(f"ERROR No valid scans found in {filename}")
		else:
			show_error("ERROR", "Unable to load file!\nMake sure file ends in .csv or .xlsx")
	except:
		show_error("ERROR", "Unable to load file!\nIs it open in another program?")
		activity_log(f"ERROR loading {database_file_entrybox.get()} Misc. error")
		log_error()

def parse_new_entries(header_line, data_in):
	#Sets all indexes to -1, which represents "index not found" later on
	OrderId_index = ReferenceNum_index = BatchOrderId_index = CreationDate_index = Carrier_index = ShipService_index = ProcessDate_index = PickTicketPrintDate_index = TrackingNumber_index = -1

	#Figure out which columns correspond to which data
	for i, cell in enumerate(header_line):
		#Ugly and hard-coded but at least it won't change (Header names are from 3PLC)
		#1st string is XLSX Header, 2nd string is CSV Header
		if cell == "Transaction ID" or cell == "OrderId": OrderId_index = i
		if cell == "Customer" : Customer_index = i
		if cell == "Reference Number" or cell == "ReferenceNum": ReferenceNum_index = i
		if cell == "Batch ID" or cell == "BatchOrderId": BatchOrderId_index = i
		if cell == "Create Date" or cell == "CreationDate": CreationDate_index = i
		if cell == "Carrier" or cell == "Carrier": Carrier_index = i
		if cell == "Ship Service" or cell == "ShipService": ShipService_index = i
		if cell == "Close Date" or cell == "ProcessDate": ProcessDate_index = i
		if cell == "Pick Ticket Print Date" or cell == "PickTicketPrintDate": PickTicketPrintDate_index = i #TODO Shrink
		if cell == "Tracking Number" or cell == "TrackingNumber": TrackingNumber_index = i
		
	if all((OrderId_index != -1, ProcessDate_index != -1, TrackingNumber_index != -1)): #If an Order ID is found (used as a unique ID for entries, absolutely necesssary)
		global database_data #Where the data is appended to
		for row in data_in:	#Build the output row to append to database
			out_row = []
			#Loops through every index and appends any data that matched
			for i in (ReferenceNum_index, Customer_index, OrderId_index, BatchOrderId_index, CreationDate_index, Carrier_index, ShipService_index, ProcessDate_index, PickTicketPrintDate_index, TrackingNumber_index):
				#String-quotes tracking numbers
				if i == TrackingNumber_index:
					out_row.append("'" + str(row[i]) if i >= 0 else "N/A")
				else:
					out_row.append(str(row[i]) if i >= 0 else "N/A") #i = -1 if index not found\

			''' New Policy is to add everything to database, and mark status accordingly
			#check if row is valid
			if is_valid_database_row(out_row): 
				database_data.append(out_row)
			'''
			out_row = set_untrackable_entries_as_shipped(out_row)
			database_data.append(out_row)

		database_data = remove_duplicates(database_data)

		show_info("Success!", "New orders matched, duplicates removed.\n\nPlease wait while we update the database. . .")
		activity_log(f"SUCCESS loading {database_file_entrybox.get()}")
		save_database(database_data)
	else:
		show_error("ERROR", "One of \"Transaction ID\", \"ProcessDate\", or \"Tracking Number\" not found in first row of uploaded file!\nAdjust your 3PLC columns to include those headers!")
		activity_log(f"ERROR loading {database_file_entrybox.get()}: File lacks \"Transaction ID\", \"ProcessDate\", or \"Tracking Number\"")

def remove_duplicates(data_in):
	#Makes a list of unique order numbers
	#Then loops through that list, matching order numbers, and merging entries where appropriate

	#Make list of unique orders
	unique_orders = []
	for row in data_in:
		order_id = row[2]
		if order_id not in unique_orders:
			unique_orders.append(order_id)

	data_out = []
	for order_number in unique_orders:
		line_out = ["", "", "", "", "", "", "", "", "", "", "", ""]
		for row in data_in:
			order_id = row[2]
			if order_number == order_id:
				line_out = merge_row(line_out, right_pad_list(row, 12, ""))
		data_out.append(line_out)
	return data_out

def merge_row(base_row, add_row):
    #print(base_row, add_row)
    row_out = []
    for i in range(len(base_row)):
    	if base_row[i] == "" or base_row[i] == "None":
    		row_out.append(add_row[i])
    	else:
	        if add_row[i] > base_row[i]:
	            row_out.append(add_row[i])
	        else:
	            row_out.append(base_row[i])
    return row_out

def set_untrackable_entries_as_shipped(row_in):
	'''Identifies packages that can't be tracked/scanned and sets their status accordingly'''
	global config_data
	row_out = row_in
	if row_in[5] in config_data["valid_carriers"]: 
		return row_out
	else:
		#UPS
		if row_in[9][0:3] in config_data["ups_tracking_number_info"]["start_digits"] and len(row_in[9]) == config_data["ups_tracking_number_info"]["length"] + 1:
			pass
		#Canada Post
		elif row_in[9][0:3] in config_data["canadapost_tracking_number_info"]["start_digits"] and len(row_in[9]) == config_data["canadapost_tracking_number_info"]["length"] + 1:
			pass
		#EHub
		elif row_in[9][0:3] in config_data["usps_tracking_number_info"]["start_digits"] and len(row_in[9]) == config_data["usps_tracking_number_info"]["length"] + 1:
			pass
		#DHL
		elif len(row_in[9]) == config_data["dhl_tracking_number_info"]["length"] + 1:
			pass
		#Lettermail
		elif fuzzy_match(row_in[9], "LETTERMAIL"):
			row_out.append("LETTERMAIL")
			row_out.append(row_in[4][0:10]) #TODO Try:catch
		#Cancelled Orders
		elif "CANCEL" in row_in[0]:
			row_out.append("CANCELLED")
			row_out.append(row_in[4][0:10])
		#Retail Orders
		elif "#VM" in row_in[0]:
			row_out.append("RETAIL")
			row_out.append(row_in[4][0:10])
		else:
			row_out.append("MISC")
			row_out.append(row_in[4][0:10])

	return row_out

def is_valid_database_row(row):
	if is_invalid_value(row[1]): #CustomerID
		return False
	if is_invalid_value(row[2]): #OrderID
		return False
	if is_invalid_value(row[4]): #Creation Date
		return False
	if is_invalid_value(row[5]): #Carrier
		return False
	if is_invalid_value(row[6]): #Ship Service. Entries lacking Ship Service can still be valid
		if is_invalid_value(row[5]) or is_invalid_value(row[9]): #If the carrier isn't NLS and the tracking number isn't invalid
			return False
	if is_invalid_value(row[9]) or fuzzy_match(row[9], "None"): #Tracking Number
		return False
	#If no reason to reject line is found
	return True

def is_invalid_value(value):
	global config_data
	if fuzzy_match(value, config_data["rejection_values"]) or value == None or value == "":
		return True
	else:
		return False

def right_pad_list(data, length, filler):
	_out = data
	for i in range(length - len(data)):
		_out.append(filler)
	return _out

## Upload to Google Sheets Functions

def upload_to_google_docs():
	try:
		global database_data
		global config_data
		for client in config_data["clients_list"]:
			client_data = filter_by_client(database_data, client)
			client_sheet = get_client_spreadsheet_id(client)
			#for i, row in enumerate(client_data):
			#	if i < 3:
			#		print(row)
			result = gapi.main(in_spreadsheet_id = client_sheet, in_range = "MAIN!A:A", data_in = client_data)
			messagebox.showinfo("MAYBE?", "Google Sheet Updated!\nCheck Activity log (black window) and Google Sheet to confirm!\n(Will update this message once I get a better sample of errors/error messages)")
			# TODO Return Status Parser
			activity_log(f"GAPI {result}")
	except: 
		log_error()

def get_client_spreadsheet_id(client):
	#Do a client -> Google Sheet.ID lookup in config.json
	for client_name, sheet_id in config_data["clients_sheets"].items():
		if client == client_name:
			return sheet_id
	show_error("ERROR", "Client spreadsheet not specified in CONFIG.json")
	return ""

def filter_by_client(data, client):
	global config_data
	out_data = []
	out_data.append(data[0]) #Add the header
	for row in data:
		if row[config_data["database_Client_index"]] == client:
			out_data.append(row)
	#print(out_data)
	return out_data
'''
def get_and_filter_database_by_date(start_date, end_date):
	global database_data
	load_database()
	out_data = []
	out_data.append(database_data[0]) #Add the header
	for row in database_data:
		if row[config_data["database_Client_index"]] in config_data["clients_list"]:
			if row[config_data["database_row_length"] - 2] == "SHIPPED":
				#For shipped entries, filter out anything older than the matched_entries_max_age
				if not is_older_than_x_days(row[config_data["database_row_length"] - 1][:10], matched_entries_max_age):
					out_data.append(row)
			else:
				#For unshipped entries, filter out anything older than the unmatched_entries_max_age
				if not is_older_than_x_days(row[config_data["database_CreationDate_index"]], unmatched_entries_max_age):
					out_data.append(row)
	return out_data

def is_older_than_x_days(date_in: str, days_in):
	if is_only_numeric(date_in): #If the date is actually a date
		try:
			entry_date = datetime.strptime(date_in, "%Y-%m-%d") #TODO Shrink
		except:
			try:
				entry_date = datetime.strptime(date_in, "%Y-%m-%d %H:%M:%S")
			except:
				entry_date = datetime.strptime(date_in, "%Y-%m-%d %H:%M")
		now_date = datetime.now()
		time_delta = datetime_module.timedelta(days = days_in)
		#print(entry_date, (now_date - time_delta), (entry_date < now_date - time_delta))
		if (entry_date < now_date - time_delta):
			return True
		else:
			return False
	else: return False
'''
## Report Generation

def generate_report():
	global config_data
	global database_data

	try:
		start_date = datetime.strptime(start_date_entrybox.get(), "%Y-%m-%d")
		end_date = datetime.strptime(end_date_entrybox.get(), "%Y-%m-%d")

		print(start_date, end_date)

		out_data = []
		for i, row in enumerate(database_data):
			if row[10] != "SHIPPED" and i != 0:
				try:
					row_date = datetime.strptime(row[7], "%Y-%m-%d %H:%M:%S")
				except:
					row_date = datetime.strptime(row[7], "%Y-%m-%d %H:%M")
				if row_date >= start_date and row_date <= end_date:
					out_data.append(row)

		if len(out_data) > 0:
			filename = "Unshipped_Report_" + datetime.strftime(datetime.today(), "%Y-%m-%d %H.%M.%S") + ".csv"
			with open(filename, "w", newline = "", encoding="utf-8") as file:
				csv_writer = csv.writer(file)
				for row in out_data:
					csv_writer.writerow(row)
			show_info("SUCCESS", f"Unshipped Report successfully generated!\nPlease see {filename}")
		else:
			show_error("ERROR", "No entries found for specified date range!")

	except:
		show_error("ERROR", "Unable to parse dates!\nMake sure dates are in YYYY-MM-DD format!")
		traceback.print_exc()

## Utility Functions

def is_not_empty(list_in):
	not_empty = False
	for entry in list_in:
		if entry:
			not_empty = True
	return not_empty

def fuzzy_match(value, target):
	if type(target) == list and type(target) != str:
		result = False
		for element in target:
			if fuzzy_match(value, element):
				result = True
		return result
	else:
		return (to_lower_alpha(value) == to_lower_alpha(target))

def is_only_numeric(string_in):
	for c in string_in:
		if c not in "1234567890 -:":
			return False
	return True

def to_lower_alpha(string_in):
	return "".join(c for c in string_in if c in "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz").lower()

def exit_program(*keys):
	print("\nExiting . . .")
	sys.exit(0)

def log_error():
	if not os.path.exists("resources/Error_Log.txt"):
		file = open("resources/Error_Log.txt", "w")
		file.write("Error_Log")
		file.close()

	traceback.print_exc()
	with open("resources/Error_Log.txt", "w", encoding = "utf-8") as file:
		file.write("\n")
		traceback.print_exc(file = file)

def show_error(title, text):
	global verbose_flag
	if verbose_flag:
		messagebox.showerror(title, text)
	else:
		print(title, text)

def show_info(title, text):
	global verbose_flag
	if verbose_flag:
		messagebox.showinfo(title, text)
	else:
		print(title, text)

def toggle_verbose():
	global verbose_flag
	verbose_flag = not verbose_flag
	if verbose_flag:
		print("Displaying Message Boxes")
	else:
		print("Hiding Message Boxes")

def set_verbose(boolean = False):
	global verbose_flag
	verbose_flag = boolean
	if verbose_flag:
		print("Displaying Message Boxes")
	else:
		print("Hiding Message Boxes")

### Global Variables
config_data = {}
database_data = []
header_data_line = []
verbose_flag = True

### Main
config_data = load_config()
program_title = config_data.get("program_name") + " " + config_data.get("program_version")
print(program_title)

## tkInterface setup
main_window = tk.Tk()
main_window.title(program_title)

main_window.protocol("WM_DELETE_WINDOW", exit_program)
main_window.bind('<Escape>', exit_program)

#Menu bar
menu_bar = tk.Menu(main_window)
main_window.config(menu = menu_bar)

file_menu = tk.Menu(menu_bar, tearoff = 0)
file_menu.add_command(label = "Exit", command = exit_program)
menu_bar.add_cascade(label = "Exit", menu = file_menu)

#Frames
scans_tab = tk.Frame(main_window)
scans_tab.grid(column = 0, row = 2, pady = 5, sticky = "W")

seperator_1_frame = tk.Frame(main_window)
seperator_1_frame.grid(column = 0, row = 1, pady = 5, sticky = "EW")

database_tab = tk.Frame(main_window)
database_tab.grid(column = 0, row = 0, pady = 5, sticky = "W")

seperator_2_frame = tk.Frame(main_window)
seperator_2_frame.grid(column = 0, row = 3, pady = 5, sticky = "EW")

info_tab = tk.Frame(main_window)
info_tab.grid(column = 0, row = 4, pady = 4, sticky = "W")

seperator_3_frame = tk.Frame(main_window)
seperator_3_frame.grid(column = 0, row = 5, pady = 5, sticky = "EW")

report_tab = tk.Frame(main_window)
report_tab.grid(column = 0, row = 6, pady = 5, sticky = "W")

seperator_4_frame = tk.Frame(main_window)
seperator_4_frame.grid(column = 0, row = 7, pady = 5, sticky = "EW")

utilities_frame = tk.Frame(main_window)
utilities_frame.grid(column = 0, row = 8, pady = 5, sticky = "W")

#File Entry
scans_file_label = tk.Label(scans_tab, text = "Upload Scans File:")
scans_file_label.grid(column = 0, row = 0, pady = 2, sticky = "W")

scans_file_entrybox = tk.Entry(scans_tab, width = 100)
scans_file_entrybox.grid(column = 0, row = 1, pady = 2, sticky = "W")

scans_file_entrybutton = tk.Button(scans_tab, text = "3. Browse", command = scans_browse_filesystem)
scans_file_entrybutton.grid(column = 0, row = 2, pady = 2, sticky = "W")

load_scans_button = tk.Button(scans_tab, text = "4. Load file", command = load_scans)
load_scans_button.grid(column = 0, row = 3, pady = 2, sticky = "W")

##Seperator 1
seperator_1 = ttk.Separator(seperator_1_frame, orient='horizontal')
seperator_1.place(x = 0, y = 0, relwidth = 1)

##Database Tab
file_entry_label = tk.Label(database_tab, text = "Upload file to Database:")
file_entry_label.grid(column = 0, row = 0, pady = 2, sticky = "W")

database_file_entrybox = tk.Entry(database_tab, width = 100)
database_file_entrybox.grid(column = 0, row = 1, pady = 2, sticky = "W")

database_file_entrybutton = tk.Button(database_tab, text = "1. Browse", command = database_browse_filesystem)
database_file_entrybutton.grid(column = 0, row = 2, pady = 2, sticky = "W")

load_new_entries_button = tk.Button(database_tab, text = "2. Load New Entries", command = load_new_entries)
load_new_entries_button.grid(column = 0, row = 3, pady = 2, sticky = "W")

database_status_label = tk.Label(database_tab, text = "Status: Database not loaded!")
database_status_label.grid(column = 0, row = 4, pady = 2, sticky = "W")

load_database()

##Seperator 2
seperator_2 = ttk.Separator(seperator_2_frame, orient='horizontal')
seperator_2.place(x = 0, y = 0, relwidth = 1)

##Info Tab
update_label = tk.Label(info_tab, text = "Upload Database to Google Sheets:")
update_label.grid(column = 0, row = 0, pady = 2, sticky = "W")

upload_button = tk.Button(info_tab, text = "5. Upload to Google Sheets", command = upload_to_google_docs)
upload_button.grid(column = 0, row = 1, pady = 2, sticky = "W")

##Seperator 3
seperator_3 = ttk.Separator(seperator_3_frame, orient='horizontal')
seperator_3.place(x = 0, y = 0, relwidth = 1)

##Report Tab

report_tab_label = tk.Label(report_tab, text = "Unshipped Packages Report")
report_tab_label.grid(column = 0, row = 0, pady = 2, sticky = "W")

start_date_label = tk.Label(report_tab, text = "Start ClosedDate (YYYY-MM-DD):")
start_date_label.grid(column = 0, row = 1, pady = 5, sticky = "W")

start_date_entrybox = tk.Entry(report_tab)
start_date_entrybox.insert(0, datetime.strftime(datetime.today(), "%Y-%m-%d"))
start_date_entrybox.grid(column = 1, row = 1, pady = 2, sticky = "W")

end_date_label = tk.Label(report_tab, text = "End ClosedDate (YYYY-MM-DD):")
end_date_label.grid(column = 0, row = 2, pady = 2, sticky = "W")

end_date_entrybox = tk.Entry(report_tab)
end_date_entrybox.insert(0, datetime.strftime(datetime.today(), "%Y-%m-%d"))
end_date_entrybox.grid(column = 1, row = 2, pady = 2, sticky = "W")

report_gen_button = tk.Button(report_tab, text = "Generate Report", command = generate_report)
report_gen_button.grid(column = 0, row = 3, pady = 2, sticky = "W")

##Seperator 4
seperator_4 = ttk.Separator(seperator_4_frame, orient='horizontal')
seperator_4.place(x = 0, y = 0, relwidth = 1)

## Progress Bar
utilities_label = tk.Label(utilities_frame, text = "Utilities:")
utilities_label.grid(column = 0, row = 0, pady = 2, sticky = "W")

backup_database_button = tk.Button(utilities_frame, text = "Backup Database", command = backup_database)
backup_database_button.grid(column = 0, row = 1, pady = 2, sticky = "W")

restore_database_button = tk.Button(utilities_frame, text = "Restore Database", command = restore_database)
restore_database_button.grid(column = 0, row = 2, pady = 2, sticky = "W")

clean_backups_button = tk.Button(utilities_frame, text = "Clean old Backups", command = clean_old_backups)
clean_backups_button.grid(column = 0, row = 3, pady = 2, sticky = "W")

upload_scans_folder = tk.Button(utilities_frame, text = "Upload all scans in /scans", command = upload_scans_folder)
upload_scans_folder.grid(column = 0, row = 4, pady = 2, sticky = "W")

toggle_verbose_button = tk.Button(utilities_frame, text = "Display messages toggle", command = toggle_verbose)
toggle_verbose_button.grid(column = 0, row = 5, pady = 2, sticky = "W")

scans_date_label = tk.Label(utilities_frame, text = "Scan Date Override: YYYY-MM-DD")
scans_date_label.grid(column = 0, row = 6, pady = 2, sticky = "W")

scans_date_entrybox = tk.Entry(utilities_frame, width = 100)
scans_date_entrybox.grid(column = 0, row = 7, pady = 2, sticky = "W")

##Misc
#show_info("Welcome!", "It's not frozen!\nIt's actually just processing a lot of data!")
main_window.mainloop()
sys.exit(0)